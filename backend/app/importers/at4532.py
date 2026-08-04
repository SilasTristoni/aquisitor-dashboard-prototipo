import io
import re
from datetime import UTC, date, datetime, time

from openpyxl import load_workbook

from app.domain.readings import TemperatureChannelReading, TemperatureReading, normalize_temperature
from app.importers.common import ImportIssue, ImportResult, ensure_aware

CHANNEL_PATTERN = re.compile(
    r"^(?:ch(?:annel)?|t|canal|termopar)\s*0?(\d{1,2})(?:\s*[\[(](.+?)[\])])?$", re.I
)


def _header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def _quality(value: object) -> tuple[float | None, str]:
    if value is None or str(value).strip() == "":
        return None, "missing"
    if isinstance(value, str) and value.strip().lower() in {
        "ol",
        "over",
        "overload",
        "----",
        "error",
        "err",
    }:
        return None, "overload" if value.strip().lower() in {
            "ol",
            "over",
            "overload",
        } else "invalid"
    try:
        return float(str(value).replace(",", ".")), "good"
    except ValueError:
        return None, "invalid"


def _json_value(value: object) -> object:
    if isinstance(value, datetime | date | time):
        return value.isoformat()
    return value


class At4532XlsxImporter:
    max_rows = 1_000_000
    max_channels = 32

    def parse(self, payload: bytes) -> ImportResult[TemperatureReading]:
        if not payload.startswith(b"PK"):
            raise ValueError("Arquivo não possui assinatura XLSX válida")
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        try:
            sheet = workbook.active
            header_row, mapping, units = self._find_header(sheet)
            result: ImportResult[TemperatureReading] = ImportResult(
                mapping={key: str(value) for key, value in mapping.items()}
            )
            for row_number, values in enumerate(
                sheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1
            ):
                if row_number - header_row > self.max_rows:
                    raise ValueError("Arquivo excede o limite de linhas")
                if not any(value is not None for value in values):
                    continue
                try:
                    timestamp = self._timestamp(values, mapping)
                    channels = []
                    for key, column in mapping.items():
                        if not key.startswith("channel_"):
                            continue
                        channel = int(key.split("_")[1])
                        original, quality = _quality(values[column])
                        unit = units.get(key, "°C")
                        canonical = (
                            normalize_temperature(original, unit) if original is not None else None
                        )
                        channels.append(
                            TemperatureChannelReading(
                                channel=channel,
                                temperature_c=canonical,
                                original_value=original,
                                original_unit=unit,
                                quality=quality,
                            )
                        )
                    ambient = None
                    if "ambient" in mapping:
                        original, _ = _quality(values[mapping["ambient"]])
                        ambient = normalize_temperature(original, units.get("ambient", "°C"))
                    result.readings.append(
                        TemperatureReading(
                            device_timestamp=timestamp,
                            received_timestamp=timestamp,
                            ambient_temperature_c=ambient,
                            channels=channels,
                            quality="good"
                            if any(item.quality == "good" for item in channels)
                            else "invalid",
                            raw_payload={
                                str(index + 1): _json_value(value)
                                for index, value in enumerate(values)
                            },
                        )
                    )
                except (ValueError, TypeError, IndexError) as exc:
                    result.errors.append(ImportIssue(row_number, "row", str(exc)))
            return result
        finally:
            workbook.close()

    def _find_header(self, sheet) -> tuple[int, dict[str, int], dict[str, str]]:
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=1, max_row=80, values_only=True), 1
        ):
            mapping: dict[str, int] = {}
            units: dict[str, str] = {}
            for index, value in enumerate(values):
                label = _header(value)
                match = CHANNEL_PATTERN.match(label)
                if match and 1 <= int(match.group(1)) <= self.max_channels:
                    key = f"channel_{int(match.group(1))}"
                    mapping[key] = index
                    if match.group(2):
                        units[key] = match.group(2).strip()
                elif label in {"timestamp", "datetime", "data hora", "date time"}:
                    mapping["timestamp"] = index
                elif label in {"data", "date"}:
                    mapping["date"] = index
                elif label in {"hora", "time"}:
                    mapping["time"] = index
                elif "ambient" in label or label in {"amb", "temperatura ambiente"}:
                    mapping["ambient"] = index
                    unit_match = re.search(r"[\[(](.+?)[\])]", label)
                    if unit_match:
                        units["ambient"] = unit_match.group(1).strip()
            has_time = "timestamp" in mapping or "time" in mapping
            if has_time and any(key.startswith("channel_") for key in mapping):
                return row_number, mapping, units
        raise ValueError("Cabeçalho do AT4532 não reconhecido")

    def _timestamp(self, row: tuple, mapping: dict[str, int]) -> datetime:
        if "timestamp" in mapping:
            value = row[mapping["timestamp"]]
            if isinstance(value, datetime):
                return ensure_aware(value)
            return ensure_aware(datetime.fromisoformat(str(value).replace("Z", "+00:00")))
        date_value = row[mapping["date"]] if "date" in mapping else date.today()
        time_value = row[mapping["time"]]
        if isinstance(date_value, datetime):
            date_value = date_value.date()
        elif not isinstance(date_value, date):
            date_value = datetime.strptime(str(date_value), "%d/%m/%Y").date()
        if isinstance(time_value, datetime):
            time_value = time_value.time()
        elif not isinstance(time_value, time):
            time_value = time.fromisoformat(str(time_value))
        return datetime.combine(date_value, time_value, tzinfo=UTC)
