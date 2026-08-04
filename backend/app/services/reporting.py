import csv
import io
from datetime import UTC, datetime

from openpyxl import Workbook
from PIL import Image, ImageDraw
from reportlab.graphics.charts.lineplots import LinePlot
from reportlab.graphics.shapes import Drawing
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.models.entities import AlertEvent, Measurement, MeasurementSession, Report
from app.services.statistics import session_statistics
from app.services.synchronization import synchronized_series

HEADERS = [
    "timestamp_grade",
    "timestamp_temperatura",
    "timestamp_eletrica",
    "tensao_v",
    "corrente_a",
    "potencia_ativa_w",
    "potencia_aparente_va",
    "potencia_reativa_var",
    "fator_potencia",
    "frequencia_tensao_hz",
    "frequencia_corrente_hz",
    "temperatura_ambiente_c",
] + [f"temperatura_t{channel}_c" for channel in range(1, 33)]


def _session_and_rows(db: Session, session_id: int) -> tuple[MeasurementSession, list[Measurement]]:
    session = db.get(MeasurementSession, session_id)
    if not session:
        raise ValueError("Sessão não encontrada")
    rows = list(
        db.scalars(
            select(Measurement)
            .options(selectinload(Measurement.temperatures))
            .where(Measurement.session_id == session_id)
            .order_by(Measurement.timestamp)
        )
    )
    return session, rows


def _row(measurement: Measurement) -> list:
    values = {temp.channel: temp.temperature_c for temp in measurement.temperatures}
    return [
        measurement.timestamp.isoformat(),
        measurement.raw_power,
        measurement.raw_power_unit,
        measurement.power_w,
        *[values.get(channel) for channel in range(1, 33)],
    ]


def _export_rows(db: Session, session: MeasurementSession) -> list[list]:
    synchronized = synchronized_series(
        db,
        session.id,
        grid_ms=session.sync_grid_ms,
        tolerance_ms=session.sync_tolerance_ms,
        max_points=100_000,
    )
    if synchronized["points"]:
        return [
            [
                point["timestamp"],
                point["temperature_sample_timestamp"],
                point["electrical_sample_timestamp"],
                point["voltage_v"],
                point["current_a"],
                point["active_power_w"],
                point["apparent_power_va"],
                point["reactive_power_var"],
                point["power_factor"],
                point["voltage_frequency_hz"],
                point["current_frequency_hz"],
                point["ambient_temperature_c"],
                *[point["temperatures_c"].get(str(channel)) for channel in range(1, 33)],
            ]
            for point in synchronized["points"]
        ]
    _, measurements = _session_and_rows(db, session.id)
    return [
        [
            measurement.timestamp.isoformat(),
            measurement.timestamp.isoformat(),
            measurement.timestamp.isoformat(),
            None,
            None,
            measurement.power_w,
            None,
            None,
            None,
            None,
            None,
            None,
            *[
                {value.channel: value.temperature_c for value in measurement.temperatures}.get(
                    channel
                )
                for channel in range(1, 33)
            ],
        ]
        for measurement in measurements
    ]


def create_csv(db: Session, session_id: int, user_id: int) -> bytes:
    session, _ = _session_and_rows(db, session_id)
    output = io.StringIO(newline="")
    writer = csv.writer(output, delimiter=";")
    writer.writerow(HEADERS)
    writer.writerows(_export_rows(db, session))
    db.add(Report(session_id=session_id, type="csv", generated_by=user_id))
    db.commit()
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def create_xlsx(db: Session, session_id: int, user_id: int) -> bytes:
    session, _ = _session_and_rows(db, session_id)
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Medições")
    sheet.append(HEADERS)
    for row in _export_rows(db, session):
        sheet.append(row)
    summary = workbook.create_sheet("Resumo")
    summary.append(["ThermoPower Monitor", session.name])
    stats = session_statistics(db, session_id)
    summary.append(["Amostras", stats["power"]["count"]])
    summary.append(["Potência média (W)", stats["power"]["mean"]])
    summary.append(["Potência máxima (W)", stats["power"]["max"]])
    summary.append(["Alertas", stats["alert_count"]])
    stream = io.BytesIO()
    workbook.save(stream)
    db.add(Report(session_id=session_id, type="xlsx", generated_by=user_id))
    db.commit()
    return stream.getvalue()


def create_pdf(db: Session, session_id: int, user_id: int, orientation: str = "landscape") -> bytes:
    session, measurements = _session_and_rows(db, session_id)
    stats = session_statistics(db, session_id)
    output = io.BytesIO()
    page_size = landscape(A4) if orientation == "landscape" else A4
    document = SimpleDocTemplate(
        output,
        pagesize=page_size,
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("<b>ThermoPower Monitor</b>", styles["Title"]),
        Paragraph(f"Relatório da sessão: {session.name}", styles["Heading2"]),
        Paragraph(
            f"Gerado em {datetime.now(UTC).strftime('%d/%m/%Y %H:%M UTC')}", styles["Normal"]
        ),
        Paragraph(
            f"Equipamento: {session.device.name} &nbsp;&nbsp; Operador: {session.user.name}",
            styles["Normal"],
        ),
        Paragraph(f"Observações: {session.notes or 'Não informadas'}", styles["Normal"]),
        Spacer(1, 6 * mm),
    ]
    summary_data = [
        ["Indicador", "Valor"],
        ["Amostras", str(stats["power"]["count"])],
        ["Potência média", f"{(stats['power']['mean'] or 0):.2f} W"],
        ["Potência mínima", f"{(stats['power']['min'] or 0):.2f} W"],
        ["Potência máxima", f"{(stats['power']['max'] or 0):.2f} W"],
        ["Alertas", str(stats["alert_count"])],
        ["Lacunas detectadas", str(stats["acquisition_gaps"])],
    ]
    summary_table = Table(summary_data, colWidths=[55 * mm, 45 * mm])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17233F")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("PADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    story.extend([summary_table, Spacer(1, 6 * mm)])
    if len(measurements) > 1:
        stride = max(1, len(measurements) // 120)
        chart_values = [
            (index, measurement.power_w) for index, measurement in enumerate(measurements[::stride])
        ]
        drawing = Drawing(240 * mm, 55 * mm)
        plot = LinePlot()
        plot.x, plot.y, plot.width, plot.height = 12 * mm, 8 * mm, 210 * mm, 40 * mm
        plot.data = [chart_values]
        plot.lines[0].strokeColor = colors.HexColor("#3569ED")
        plot.lines[0].strokeWidth = 1.5
        plot.xValueAxis.visibleLabels = False
        plot.yValueAxis.labelTextFormat = "%0.0f W"
        drawing.add(plot)
        story.extend([Paragraph("Curva de potência", styles["Heading3"]), drawing])
    alerts = list(
        db.scalars(
            select(AlertEvent)
            .where(AlertEvent.session_id == session_id)
            .order_by(AlertEvent.timestamp)
            .limit(15)
        )
    )
    if alerts:
        alert_data = [["Horário", "Métrica", "Canal", "Valor", "Limite", "Severidade"]]
        alert_data.extend(
            [
                alert.timestamp.strftime("%H:%M:%S"),
                alert.metric,
                f"T{alert.channel}" if alert.channel else "—",
                f"{alert.measured_value:.2f}",
                f"{alert.threshold:.2f}",
                alert.severity,
            ]
            for alert in alerts
        )
        alert_table = Table(alert_data, repeatRows=1)
        alert_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFF0F1")),
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                ]
            )
        )
        story.extend([Paragraph("Alertas", styles["Heading3"]), alert_table, Spacer(1, 5 * mm)])
    story.append(Paragraph("Amostra das medições", styles["Heading3"]))
    sample_headers = ["Horário", "Original", "Potência (W)", "T1", "T2", "T3", "T4"]
    sample_data = [sample_headers]
    for measurement in measurements[:40]:
        row = _row(measurement)
        sample_data.append([row[0][11:19], f"{row[1]} {row[2]}", f"{row[3]:.2f}", *row[4:8]])
    sample_table = Table(sample_data, repeatRows=1)
    sample_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EEFF")),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("PADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(sample_table)
    document.build(story)
    db.add(Report(session_id=session_id, type="pdf", generated_by=user_id))
    db.commit()
    return output.getvalue()


def create_chart_image(db: Session, session_id: int, user_id: int, image_type: str) -> bytes:
    session, _ = _session_and_rows(db, session_id)
    rows = _export_rows(db, session)
    width, height = 1920, 1080
    image = Image.new("RGB", (width, height), "#f7f9fd")
    draw = ImageDraw.Draw(image)
    left, top, right, bottom = 130, 110, width - 130, height - 150
    draw.text((left, 45), f"ThermoPower Monitor — {session.name}", fill="#17233f")
    draw.rectangle((left, top, right, bottom), outline="#cbd5e1", width=2)
    power_values = [row[5] for row in rows if row[5] is not None]
    temperature_values = [
        sum(values) / len(values)
        for row in rows
        if (values := [value for value in row[12:] if value is not None])
    ]
    if rows and (power_values or temperature_values):
        power_min, power_max = _bounds(power_values)
        temp_min, temp_max = _bounds(temperature_values)
        power_points = []
        temp_points = []
        for index, row in enumerate(rows):
            x = left + (right - left) * index / max(len(rows) - 1, 1)
            if row[5] is not None:
                y = bottom - (row[5] - power_min) / (power_max - power_min) * (bottom - top)
                power_points.append((x, y))
            values = [value for value in row[12:] if value is not None]
            if values:
                average = sum(values) / len(values)
                y = bottom - (average - temp_min) / (temp_max - temp_min) * (bottom - top)
                temp_points.append((x, y))
        if len(power_points) > 1:
            draw.line(power_points, fill="#3569ed", width=4)
        if len(temp_points) > 1:
            draw.line(temp_points, fill="#f59e0b", width=4)
        draw.text(
            (left, bottom + 30),
            f"Potência ativa: {power_min:.2f}–{power_max:.2f} W",
            fill="#3569ed",
        )
        draw.text(
            (left + 430, bottom + 30),
            f"Temperatura média: {temp_min:.2f}–{temp_max:.2f} °C",
            fill="#b86a00",
        )
    else:
        draw.text((left + 30, top + 30), "Sessão sem pontos para renderizar", fill="#64748b")
    output = io.BytesIO()
    format_name = "JPEG" if image_type in {"jpg", "jpeg"} else "PNG"
    image.save(output, format=format_name, quality=95 if format_name == "JPEG" else None)
    db.add(Report(session_id=session_id, type=image_type, generated_by=user_id))
    db.commit()
    return output.getvalue()


def _bounds(values: list[float]) -> tuple[float, float]:
    if not values:
        return 0.0, 1.0
    minimum, maximum = min(values), max(values)
    if minimum == maximum:
        return minimum - 0.5, maximum + 0.5
    return minimum, maximum
